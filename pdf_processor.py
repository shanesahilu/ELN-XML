import xml.etree.ElementTree as ET
import base64
import io
try:
    import openpyxl
except ImportError:
    print("ERR: openpyxl missing. Install: <your_python_path> -m pip install openpyxl")
    exit() 

try:
    import docx
except ImportError:
    print("WARN: python-docx library not found. Word document extraction (documentType='1') will be skipped. Install with: pip install python-docx")
    docx = None

from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Preformatted
from reportlab.platypus import Image
from reportlab.lib.utils import ImageReader
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import html
import os 
import re
import sys 

LOGO_PATH = None

def set_pdf_logo_path(path):
    global LOGO_PATH
    LOGO_PATH = path

HIERARCHY_TABLE_COLUMN_MAPPINGS = {}
styles = getSampleStyleSheet()
style_h1 = styles['h1']
style_h2 = ParagraphStyle(name='Heading2', parent=styles['h2'], spaceBefore=10, spaceAfter=5, fontSize=14, fontName='Helvetica-Bold')
style_h3 = ParagraphStyle(name='Heading3', parent=styles['h3'], spaceBefore=8, spaceAfter=3, leftIndent=0.1*inch, fontSize=11, fontName='Helvetica-Bold')
style_h4_table_title = ParagraphStyle(name='Heading4TableTitle', parent=styles['h4'], spaceBefore=6, spaceAfter=2, leftIndent=0.2*inch, fontSize=9, fontName='Helvetica-Oblique')
style_body = styles['Normal']
style_body_small = ParagraphStyle(name='BodySmall', parent=styles['Normal'], fontSize=8, leading=9)
style_code = ParagraphStyle(name='Code', parent=styles['Normal'], fontName='Courier', fontSize=7, leading=8, backColor=colors.HexColor(0xf0f0f0), borderColor=colors.lightgrey, borderWidth=0.5, borderPadding=3, spaceBefore=3, spaceAfter=3, leftIndent=0.1*inch)
style_error = ParagraphStyle(name='Error', parent=styles['Normal'], textColor=colors.red)
style_note = ParagraphStyle(name='Note', parent=styles['Normal'], fontSize=8, leading=9, textColor=colors.darkblue, fontName='Helvetica-Oblique')

def _find_actual_content_start_column(header_row_tuple, data_row_tuples_block, max_cols_to_check):
    """
    Finds the first column index (0-based) that contains actual content
    by checking the header and a sample of data rows.

    Args:
        header_row_tuple: The identified header row (tuple of cell values).
        data_row_tuples_block: A list of data row tuples.
        max_cols_to_check: The number of columns to consider (width of the table).

    Returns:
        The 0-based index of the first column with content, or 0 if table is empty/no content found.
    """
    min_content_col_idx = max_cols_to_check 

    if header_row_tuple:
        for col_idx, cell_val in enumerate(header_row_tuple[:max_cols_to_check]):
            if str(cell_val if cell_val is not None else "").strip():
                min_content_col_idx = col_idx
                break 

    for data_row_tuple in data_row_tuples_block[:5]: 
        if min_content_col_idx == 0: 
            break
        for col_idx, cell_val in enumerate(data_row_tuple[:max_cols_to_check]):
            if str(cell_val if cell_val is not None else "").strip():
                min_content_col_idx = min(min_content_col_idx, col_idx) 
                if col_idx == 0 : break 
        if min_content_col_idx == 0: break 

    return min_content_col_idx if min_content_col_idx < max_cols_to_check else 0

def _add_logo_on_every_page(canvas, doc):
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        canvas.saveState()
        try:
            desired_logo_width = 0.85 * inch 
            img_reader = ImageReader(LOGO_PATH)
            orig_width, orig_height = img_reader.getSize()
            if orig_width == 0 or orig_height == 0:
                print(f"WARN: Logo at '{LOGO_PATH}' has invalid dimensions (0 width or height).")
                canvas.restoreState()
                return
            aspect_ratio = float(orig_height) / float(orig_width)
            img_width = desired_logo_width
            img_height = desired_logo_width * aspect_ratio
            padding_from_physical_page_edge = 0.2 * inch 
            page_width, page_height = doc.pagesize
            x_position = page_width - padding_from_physical_page_edge - img_width
            y_position = page_height - padding_from_physical_page_edge - img_height 
            canvas.drawImage(LOGO_PATH, x_position, y_position,
                             width=img_width, height=img_height, mask='auto')
        except Exception as e:
            print(f"WARN: Could not draw logo '{LOGO_PATH}' on PDF page: {e}")
        canvas.restoreState()

def _is_empty_row(row_str_values):
    """
    Checks if a list of string cell values represents an empty row.
    A row is empty if all its cell values are empty strings (after stripping).
    """
    return all(not val for val in row_str_values)

def _process_active_table_block(row_block, sheet_name_for_default_title, table_index_in_sheet):
    """
    Processes a block of raw rows to identify a table title, headers, and data rows,
    including trimming of common leading empty columns.
    """
    if not row_block:
        return None

    header_row_idx = -1
    raw_header_tuple_for_processing = None 
    data_block_start_idx = 0 

    for i, current_row_tuple in enumerate(row_block):
        current_row_str_values = [str(cell if cell is not None else "").strip() for cell in current_row_tuple]
        if _is_empty_row(current_row_str_values): 
            if header_row_idx == -1:  
                data_block_start_idx = i + 1 
            continue 

        non_empty_cell_count = sum(1 for cell_val in current_row_str_values if cell_val)

        if non_empty_cell_count > 1 or (non_empty_cell_count == 1 and header_row_idx == -1):
            num_potential_text_headers = sum(1 for cell_val in current_row_str_values 
                                             if cell_val and not cell_val.replace('.', '', 1).isdigit())

            if (non_empty_cell_count > 0 and num_potential_text_headers / non_empty_cell_count >= 0.5) or header_row_idx == -1:
                header_row_idx = i
                raw_header_tuple_for_processing = current_row_tuple
                data_block_start_idx = i + 1
                break 

    if header_row_idx == -1: 
        return None

    potential_title_str = ""
    if header_row_idx > 0:
        for i in range(header_row_idx - 1, -1, -1): 
            title_row_tuple = row_block[i]

            title_row_parts = [str(cell if cell is not None else "").strip() for cell in title_row_tuple]
            if not _is_empty_row(title_row_parts): 
                potential_title_str = " ".join(filter(None, title_row_parts)) 
                break
    table_title = potential_title_str if potential_title_str else f"Table {table_index_in_sheet + 1}"

    max_cols_from_header = 0
    if raw_header_tuple_for_processing:
        for col_idx, cell_val in reversed(list(enumerate(raw_header_tuple_for_processing))):
            if str(cell_val if cell_val is not None else "").strip():
                max_cols_from_header = col_idx + 1
                break

    max_cols_from_data = 0
    for r_idx in range(data_block_start_idx, len(row_block)):
        current_data_row_tuple = row_block[r_idx]
        current_max_col_this_row = 0
        for col_idx, cell_val in reversed(list(enumerate(current_data_row_tuple))):
            if str(cell_val if cell_val is not None else "").strip():
                current_max_col_this_row = col_idx + 1
                break
        max_cols_from_data = max(max_cols_from_data, current_max_col_this_row)

    overall_max_cols = max(max_cols_from_header, max_cols_from_data)
    if overall_max_cols == 0: 
        return None

    header_for_trim_check = raw_header_tuple_for_processing if raw_header_tuple_for_processing else [None] * overall_max_cols
    data_for_trim_check = row_block[data_block_start_idx:]

    actual_content_start_col = _find_actual_content_start_column(
        header_for_trim_check,
        data_for_trim_check,
        overall_max_cols
    )

    if raw_header_tuple_for_processing:
        trimmed_header_cells_raw = raw_header_tuple_for_processing[actual_content_start_col : overall_max_cols]
        final_headers = [str(cell if cell is not None else "") for cell in trimmed_header_cells_raw]
    else: 
        num_cols_after_trim = overall_max_cols - actual_content_start_col
        if num_cols_after_trim <= 0: return None 
        final_headers = [f"Col{j+1}" for j in range(num_cols_after_trim)]

    final_data_rows = []
    for r_idx in range(data_block_start_idx, len(row_block)):
        data_row_tuple = row_block[r_idx]

        trimmed_data_cells_raw = data_row_tuple[actual_content_start_col : overall_max_cols]
        current_trimmed_row_values = [str(cell if cell is not None else "") for cell in trimmed_data_cells_raw]

        if not _is_empty_row([val.strip() for val in current_trimmed_row_values]):
            final_data_rows.append(current_trimmed_row_values)

    if final_data_rows and final_headers and all(h.startswith("Col") and h[3:].isdigit() for h in final_headers):
        first_data_row_candidate = final_data_rows[0]
        if first_data_row_candidate: 
            num_text_like_cells = sum(1 for x_str in first_data_row_candidate
                                      if x_str.strip() and not x_str.replace('.', '', 1).isdigit())

            if len(final_headers) > 0 and (num_text_like_cells / len(final_headers) > 0.6):
                if final_headers != first_data_row_candidate: 
                    final_headers = first_data_row_candidate
                    final_data_rows = final_data_rows[1:] 

    if final_headers and not any(h.strip() for h in final_headers) and final_data_rows:
        if final_data_rows[0]: 
            final_headers = [f"Col{j+1}" for j in range(len(final_data_rows[0]))]

    if not final_headers and not final_data_rows:
        return None

    if not final_data_rows and all(h.startswith("Col") and h[3:].isdigit() for h in final_headers):
        return None

    return {
        "table_title": table_title.strip(),
        "headers": final_headers,
        "rows": final_data_rows
    }

def load_column_mappings_from_schema_files(schema_file_paths_dict):
    global HIERARCHY_TABLE_COLUMN_MAPPINGS
    HIERARCHY_TABLE_COLUMN_MAPPINGS = {} 
    HIERARCHY_TABLE_COLUMN_MAPPINGS["Media Equilibration and Readiness for Vial Thaw"] = {
        "F_95": "MFSR Name", "F_96": "CO2 Incubator ID/ Water Bath ID", "F_97": "Set Temp oC",
        "F_98": "Displayed Temp oC", "F_99": "Set CO2(%)", "F_100": "Displayed CO2(%)",
        "F_101": "Set Relative Humidity (%)", "F_102": "Displayed Relative Humidity (%)",
        "F_103": "Set Agitation (RPM)", "F_104": "Displayed agitation (RPM)",
        "F_105": "Volume of Media(mL)", "F_106": "Incubation St time",
        "F_107": "Incubation End Time", "F_108": "Incubation Duration",
    }
    if schema_file_paths_dict: 
        for schema_name, file_path in schema_file_paths_dict.items():
            try:
                if not os.path.exists(file_path):
                    print(f"WARN: Schema file not found: {file_path} for {schema_name}. Skipping.")
                    continue
                tree = ET.parse(file_path)
                root = tree.getroot()
                for protocol_version in root.findall(".//protocolVersion"):
                    for table_elem in protocol_version.findall("table"):
                        table_name_attr = table_elem.get("name")
                        if not table_name_attr: continue
                        if table_name_attr not in HIERARCHY_TABLE_COLUMN_MAPPINGS:
                            HIERARCHY_TABLE_COLUMN_MAPPINGS[table_name_attr] = {}
                        for field_elem in table_elem.findall("field"):
                            field_key = field_elem.get("key")
                            field_display_name = field_elem.get("name")
                            if field_key and field_display_name:
                                internal_field_tag = f"F_{field_key}"
                                HIERARCHY_TABLE_COLUMN_MAPPINGS[table_name_attr][internal_field_tag] = field_display_name
            except ET.ParseError as e: print(f"ERR: Parsing schema XML {file_path}: {e}")
            except Exception as e: print(f"ERR: Processing schema file {file_path}: {e}")

def extract_metadata_properties(property_instances_element):
    data = []
    if property_instances_element is not None:
        for prop_instance in property_instances_element.findall("propertyInstance"):
            value = prop_instance.get("value", "")
            prop_elem = prop_instance.find("property")
            heading = prop_elem.get("name", "") if prop_elem is not None else "N/A"
            if heading.strip() or value.strip():
                data.append({'Property': heading, 'Value': value})
    return data

def extract_styled_text_content(styled_text_element):
    if styled_text_element is None: return None, "raw"
    text_tag = styled_text_element.find("text")
    if text_tag is not None and text_tag.text and text_tag.text.strip():
        return text_tag.text.strip(), "text"
    data_tag = styled_text_element.find("data")
    if data_tag is not None and data_tag.text and data_tag.text.strip():
        return data_tag.text.strip(), "rtf" 
    return None, "raw"

def parse_checklist_text_content(text_content):
    if not text_content: return {}
    descriptions = {}
    lines = text_content.splitlines()
    for line in lines:
        line = line.strip()
        if not line: continue
        match = re.match(r'^([A-Za-z0-9\s\(\)\.\-]+?)\s+(.+?)(?:\s*■?\s*(\[[A-Z\s/]+\])\s*■?)?$', line)
        if match:
            item_id = match.group(1).strip()
            description = match.group(2).strip()
            status_in_desc_match = re.search(r'\s*■?\s*(\[[A-Z\s/]+\])\s*■?$', description, flags=re.IGNORECASE)
            if status_in_desc_match:
                description = description[:status_in_desc_match.start()].strip()
            if item_id.lower() not in ["s. no.", "s.no", "sr. no.", "sr.no", "checks", "compliance"]: 
                descriptions[item_id] = description
            continue 
        if '■' in line:
            parts = line.split('■')
            if len(parts) >= 2:
                item_id = parts[0].strip()
                description_full = "■".join(parts[1:]).strip() 
                description = re.sub(r'\s*\[[A-Z\s/]+\]\s*■?$', '', description_full, flags=re.IGNORECASE).strip()
                description = description.replace("■", " ").strip() 
                if item_id and item_id.lower() not in ["s. no.", "s.no", "sr. no.", "sr.no", "checks", "compliance"]:
                    descriptions[item_id] = description
    return descriptions

def extract_tablesection_data(tablesection_element):
    if tablesection_element is None: return [], []
    headers = [prop.get("name", f"Col{i+1}") for i, prop in enumerate(tablesection_element.findall("tableProperty/property"))]
    rows_data = []
    if tablesection_element.findall("tableRow"):
        for table_row in tablesection_element.findall("tableRow"):
            row_cells = table_row.findall("tableCell")
            num_cols_to_extract = len(headers) if headers else (len(row_cells) if row_cells else 0)
            if num_cols_to_extract == 0 and not headers: 
                first_row_for_cols = tablesection_element.find("tableRow")
                if first_row_for_cols is not None:
                    num_cols_to_extract = len(first_row_for_cols.findall("tableCell"))
            current_row_values = ["" for _ in range(num_cols_to_extract)] 
            for i in range(num_cols_to_extract):
                if i < len(row_cells): 
                    current_row_values[i] = row_cells[i].get("value", "")
            rows_data.append(current_row_values)
    if not headers and rows_data and rows_data[0]: 
        headers = [f"Col{i+1}" for i in range(len(rows_data[0]))]
    return headers, rows_data

def extract_hierarchy_data_tables(hierarchy_data_element):
    tables_content = []
    if hierarchy_data_element is None: return tables_content
    for table_elem in hierarchy_data_element.findall("table"):
        table_name_xml = table_elem.get("name", "Unnamed Hierarchy Table")
        table_specific_mappings = HIERARCHY_TABLE_COLUMN_MAPPINGS.get(table_name_xml, {})
        internal_header_tags_from_data_ordered = []
        header_tags_set = set() 
        all_rows_elements = table_elem.findall("row")
        if all_rows_elements:
            for row_elem in all_rows_elements: 
                for cell_elem in row_elem: 
                    if cell_elem.tag not in header_tags_set: 
                        header_tags_set.add(cell_elem.tag)
                        internal_header_tags_from_data_ordered.append(cell_elem.tag)
        final_internal_header_tags = [
            tag for tag in internal_header_tags_from_data_ordered if tag in table_specific_mappings
        ]
        if not final_internal_header_tags and not all_rows_elements and table_specific_mappings:
            final_internal_header_tags = [tag for tag in table_specific_mappings.keys() if tag.startswith("F_")] 
            final_internal_header_tags.sort(key=lambda x: int(x.split('_')[1]) if len(x.split('_')) > 1 and x.split('_')[1].isdigit() else 9999)
        if not final_internal_header_tags: continue 
        display_headers = [table_specific_mappings[tag] for tag in final_internal_header_tags]
        current_table_rows = []
        if all_rows_elements:
            for row_elem in all_rows_elements:
                row_values = []
                for ht in final_internal_header_tags: 
                    cell = row_elem.find(ht)
                    row_values.append(cell.text.strip() if cell is not None and cell.text else "")
                current_table_rows.append(row_values)
        if display_headers: 
            tables_content.append({
                "name": table_name_xml, "headers": display_headers, "rows": current_table_rows
            })
    return tables_content

def extract_excel_data_from_base64(base64_string):
    all_sheets_data_container = [] 
    try:
        excel_binary_data = base64.b64decode(base64_string)
        excel_file_like_object = io.BytesIO(excel_binary_data)
        workbook = openpyxl.load_workbook(excel_file_like_object, data_only=True, read_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet.max_row == 0:
                continue

            current_sheet_tables = []

            all_rows_in_sheet_raw = list(sheet.iter_rows(values_only=True)) 

            active_table_block_raw_rows = [] 
            table_counter_for_sheet = 0

            for row_idx, raw_row_tuple in enumerate(all_rows_in_sheet_raw):

                current_row_str_values_for_empty_check = [str(cell if cell is not None else "").strip() for cell in raw_row_tuple]

                if _is_empty_row(current_row_str_values_for_empty_check):

                    if active_table_block_raw_rows:
                        table_data = _process_active_table_block(
                            active_table_block_raw_rows, 
                            sheet_name, 
                            table_counter_for_sheet
                        )
                        if table_data and (table_data["headers"] or table_data["rows"]): 
                            current_sheet_tables.append(table_data)
                            table_counter_for_sheet += 1
                        active_table_block_raw_rows = [] 
                else:

                    active_table_block_raw_rows.append(raw_row_tuple)

            if active_table_block_raw_rows:
                table_data = _process_active_table_block(
                    active_table_block_raw_rows, 
                    sheet_name, 
                    table_counter_for_sheet
                )
                if table_data and (table_data["headers"] or table_data["rows"]):
                    current_sheet_tables.append(table_data)

            if current_sheet_tables:
                all_sheets_data_container.append({
                    "sheet_name": sheet_name,
                    "tables": current_sheet_tables 
                })

    except Exception as e:
        print(f"EXL ERR: Failed to extract data from Excel: {e}")

        all_sheets_data_container.append({
            "sheet_name": "Error Processing Excel File",
            "tables": [{
                "table_title": "Extraction Error",
                "headers": ["Error"],
                "rows": [[f"Could not process Excel file: {e}"]]
            }]
        })
    return all_sheets_data_container

def extract_word_text_from_base64(base64_string):
    if not docx:
        return "Skipped Word document: python-docx library not found. Please install it (`pip install python-docx`)."
    if not base64_string:
        return "Skipped Word document: No content provided."
    try:
        doc_data = base64.b64decode(base64_string)
        doc_stream = io.BytesIO(doc_data)
        document = docx.Document(doc_stream)
        full_text = [para.text for para in document.paragraphs]
        return '\n'.join(full_text)
    except Exception as e:
        print(f"ERR: Failed to extract text from Word document: {e}")
        return f"Error extracting text from Word document: {str(e)[:100]}..." 

def create_pdf_table(data_list, col_widths=None, table_style_commands=None, elements_list=None, cell_style=None, header_font_name='Helvetica-Bold', cell_font_size=7):
    if not data_list or \
       (len(data_list) == 1 and (not data_list[0] or all(not str(c).strip() for c in data_list[0]))):
        return 
    style_id_suffix = f"{id(data_list)}_{cell_font_size}" 
    if cell_style is None:
        cell_style = ParagraphStyle(name=f'CellStyle_{style_id_suffix}', parent=style_body, fontSize=cell_font_size, leading=cell_font_size + 2)
    header_cell_style = ParagraphStyle(name=f'HeaderCellStyle_{style_id_suffix}', parent=cell_style, fontName=header_font_name)
    styled_data = []
    for row_idx, row in enumerate(data_list):
        current_row_style = header_cell_style if row_idx == 0 else cell_style
        styled_row = [Paragraph(html.escape(str(cell)), current_row_style) for cell in row]
        styled_data.append(styled_row)
    if not styled_data: return 
    table = Table(styled_data, colWidths=col_widths, repeatRows=1 if len(data_list) > 1 else 0) 
    base_style = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2)
    ]
    final_style = base_style + (table_style_commands if table_style_commands else [])
    table.setStyle(TableStyle(final_style))
    if elements_list is not None:
        elements_list.append(table)
    return table 

def transpose_table(data):
    if not data: return []
    return list(map(list, zip(*data)))

def create_transposed_pdf_table(original_pdf_data, available_width, elements_list,
                                table_note_name="",
                                style_for_transposed_header_row=None, 
                                cell_font_size=7):
    if not original_pdf_data or not original_pdf_data[0] or not any(str(h).strip() for h in original_pdf_data[0]):
        return False 

    transposed_data = transpose_table(original_pdf_data)
    if not transposed_data or not transposed_data[0]: return False 

    default_transposed_header_style = [('BACKGROUND',(0,0),(0,-1),colors.lightgoldenrodyellow), 
                                       ('FONTNAME', (0,0),(0,-1), 'Helvetica-Bold')]
    combined_style = default_transposed_header_style + (style_for_transposed_header_row or [])

    table_obj = create_pdf_table( 
        transposed_data,
        col_widths=None, 
        table_style_commands=combined_style,
        elements_list=elements_list,
        cell_font_size=cell_font_size
    )
    return table_obj is not None

def optimize_table_for_display(table_name, headers, rows, available_width, elements_list,
                               default_header_style, 
                               transposed_table_new_header_row_style 
                               ):
    num_orig_cols = len(headers) if headers else 0
    if num_orig_cols == 0: 
        if rows and rows[0] and isinstance(rows[0], (list, tuple)):
            num_orig_cols = len(rows[0])
            headers = [f" " for _ in range(num_orig_cols)] 
        else:
            return False 
    pdf_data_orig = [headers] + (rows if rows else [])
    if not pdf_data_orig or \
       (len(pdf_data_orig) == 1 and (not pdf_data_orig[0] or all(not str(c).strip() for c in pdf_data_orig[0]))):
       return False
    MAX_COLS_FOR_NORMAL_LAYOUT = 25 
    rendered_something = False
    if num_orig_cols <= MAX_COLS_FOR_NORMAL_LAYOUT:

        font_size_normal = 8 
        if num_orig_cols > 18: font_size_normal = 6
        elif num_orig_cols > 10: font_size_normal = 7
        col_w = available_width / num_orig_cols if num_orig_cols > 0 else available_width
        min_col_w_abs = 0.4 * inch 
        final_col_widths_calc = [max(min_col_w_abs, col_w)]*num_orig_cols if num_orig_cols > 0 else []
        current_total_width_calc = sum(final_col_widths_calc)
        if num_orig_cols > 0 and current_total_width_calc > available_width and current_total_width_calc > 0.01: 
            scale_factor = available_width / current_total_width_calc
            final_col_widths = [w * scale_factor for w in final_col_widths_calc]
        elif num_orig_cols > 0 : 
            final_col_widths = final_col_widths_calc
        else: final_col_widths = None 
        if create_pdf_table(pdf_data_orig, col_widths=final_col_widths,
                            table_style_commands=default_header_style, elements_list=elements_list,
                            cell_font_size=font_size_normal):
            rendered_something = True
    else: 

        if create_transposed_pdf_table(
            original_pdf_data=pdf_data_orig, available_width=available_width, elements_list=elements_list,
            table_note_name=table_name, 
            style_for_transposed_header_row=transposed_table_new_header_row_style, 
            cell_font_size=7 
        ): rendered_something = True
    return rendered_something

def process_xml_to_pdf(xml_string_content, schema_files_config):
    load_column_mappings_from_schema_files(schema_files_config)
    pdf_buffer = io.BytesIO()
    total_header_reservation = 1.2 * inch 
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=total_header_reservation, 
                            bottomMargin=0.5*inch)
    final_elements_for_pdf = []
    try:
        root = ET.fromstring(xml_string_content)
    except ET.ParseError as e:
        error_message = f"XML Parsing Error: {html.escape(str(e))}"
        print(f"ERR: {error_message}")
        final_elements_for_pdf.append(Paragraph(error_message, style_error))
        try:
            doc.build(final_elements_for_pdf) 
            return pdf_buffer.getvalue()
        except Exception as build_err:
            print(f"ERR: Could not build error PDF after XML parse error: {build_err}")
            return None 
    except Exception as e: 
        error_message = f"Unexpected error during XML setup: {html.escape(str(e))}"
        print(f"ERR: {error_message}")
        final_elements_for_pdf.append(Paragraph(error_message, style_error))
        try:
            doc.build(final_elements_for_pdf)
            return pdf_buffer.getvalue()
        except Exception as build_err:
            print(f"ERR: Could not build error PDF after unexpected setup error: {build_err}")
            return None

    collection_name = html.escape(root.get("name", "N/A Collection"))
    final_elements_for_pdf.append(Paragraph(f"Report: {collection_name}", style_h1))
    ct_elem = root.find("collectionType")
    if ct_elem is not None and ct_elem.get("name"):
        final_elements_for_pdf.append(Paragraph(f"Type: {html.escape(ct_elem.get('name'))}", style_h2))
    final_elements_for_pdf.append(Spacer(1, 0.1*inch)) 
    section_set_view = root.find("sectionSetView")
    all_sections = [] 
    if section_set_view is None:
        if not final_elements_for_pdf or len(final_elements_for_pdf) <=2 : 
             final_elements_for_pdf.append(Paragraph("No 'sectionSetView' (main content) found in the XML.", style_body))
    else:
        all_sections = section_set_view.findall("section")
        for section_idx, section_elem in enumerate(all_sections):
            sec_name = html.escape(section_elem.get("name", f"Section {section_idx + 1}"))
            current_section_elements_buffer = []
            section_has_renderable_content = False
            checklist_text_descriptions_for_section = {}
            first_checklist_text_content_raw = None 
            for obj_elem_prescan in section_elem.findall("object"):
                field_elem_prescan = obj_elem_prescan.find("field")
                if field_elem_prescan is not None:
                    original_field_name_prescan = field_elem_prescan.get("name")
                    if original_field_name_prescan:
                        normalized_field_name_prescan = original_field_name_prescan.lower().strip()
                        if normalized_field_name_prescan == "checklist" or normalized_field_name_prescan == "check list":
                            if not first_checklist_text_content_raw:
                                styled_text_elem_prescan = obj_elem_prescan.find("styledText")
                                if styled_text_elem_prescan is not None:
                                    content_prescan, content_type_prescan = extract_styled_text_content(styled_text_elem_prescan)
                                    if content_prescan and content_type_prescan == "text":
                                        first_checklist_text_content_raw = content_prescan
                                        checklist_text_descriptions_for_section = parse_checklist_text_content(first_checklist_text_content_raw)
                                        break 
            for obj_idx, obj_elem in enumerate(section_elem.findall("object")):
                field_elem = obj_elem.find("field")
                if field_elem is None: continue
                original_field_name_attr = field_elem.get("name", f"Unnamed Field {obj_idx + 1}")
                normalized_field_name = original_field_name_attr.lower().strip()
                field_name_display = html.escape(original_field_name_attr)
                object_elements_buffer = []
                object_produced_renderable_content = False
                page_width_l, _ = landscape(letter); available_width_l = page_width_l - 1.0*inch 
                normalized_property_table_field_identifiers = {"metadata", "checklist", "check list", "mixing and stirring-filtration", "sop"}

                prop_instances = obj_elem.find("propertyInstances")
                if prop_instances is not None and list(prop_instances) and \
                   normalized_field_name in normalized_property_table_field_identifiers:
                    metadata_list_of_dicts = extract_metadata_properties(prop_instances)
                    table_headers = ["Property", "Value"] 
                    current_metadata_rows = [[item['Property'], item['Value']] for item in metadata_list_of_dicts] if metadata_list_of_dicts else []
                    is_checklist_type_field = (normalized_field_name == "checklist" or normalized_field_name == "check list")
                    if is_checklist_type_field and metadata_list_of_dicts:
                        table_headers = ["Checklist Item", "Status"]
                        if checklist_text_descriptions_for_section: 
                            enhanced_rows = []
                            for item_dict in metadata_list_of_dicts:
                                prop_key_original = item_dict['Property']; item_value = item_dict['Value']
                                description = checklist_text_descriptions_for_section.get(prop_key_original)
                                if not description: 
                                    prop_key_norm_pi = prop_key_original.rstrip('.').strip()
                                    for text_key, desc_val in checklist_text_descriptions_for_section.items():
                                        text_key_norm_text = text_key.rstrip('.').strip()
                                        if text_key_norm_text == prop_key_norm_pi: description = desc_val; break
                                if not description: 
                                    for text_key_raw, desc_val in checklist_text_descriptions_for_section.items():
                                        text_key_norm = text_key_raw.rstrip('.').strip(); prop_key_norm = prop_key_original.rstrip('.').strip()
                                        if (prop_key_norm.startswith(text_key_norm) or text_key_norm.startswith(prop_key_norm)) and \
                                           abs(len(prop_key_norm) - len(text_key_norm)) < 4 : 
                                            description = desc_val; break
                                formatted_property = f"{prop_key_original}: {description}" if description else prop_key_original
                                if item_value == "false": formatted_value = "Not Completed"
                                elif item_value == "true": formatted_value = "Completed"
                                elif item_value == "": formatted_value = " " 
                                else: formatted_value = item_value
                                enhanced_rows.append([formatted_property, formatted_value])
                            current_metadata_rows = enhanced_rows
                        else: 
                            temp_rows = []
                            for item_dict in metadata_list_of_dicts:
                                item_value = item_dict['Value']
                                if item_value == "false": formatted_val = "Not Completed"
                                elif item_value == "true": formatted_val = "Completed"
                                elif item_value == "": formatted_val = " "
                                else: formatted_val = item_value
                                temp_rows.append([item_dict['Property'], formatted_val])
                            current_metadata_rows = temp_rows
                    elif metadata_list_of_dicts and not is_checklist_type_field: 
                         current_metadata_rows = [[item['Property'], item['Value']] for item in metadata_list_of_dicts]
                    if optimize_table_for_display(
                        table_name=field_name_display, headers=table_headers, rows=current_metadata_rows,
                        available_width=available_width_l, elements_list=object_elements_buffer,
                        default_header_style=[('BACKGROUND',(0,0),(-1,0),colors.darkslateblue), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)],
                        transposed_table_new_header_row_style=[('BACKGROUND',(0,0),(-1,0),colors.darkslateblue), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)] 
                    ): object_produced_renderable_content = True

                styled_text_elem = obj_elem.find("styledText")
                if not object_produced_renderable_content and styled_text_elem is not None:
                    content, content_type = extract_styled_text_content(styled_text_elem)
                    if content and content_type == "text":
                        should_skip_this_styled_text = False
                        if first_checklist_text_content_raw is not None and \
                           first_checklist_text_content_raw == content and \
                           checklist_text_descriptions_for_section: 
                            should_skip_this_styled_text = True
                        if not should_skip_this_styled_text:
                            display_content = (content[:2500] + '...') if len(content) > 2500 else content
                            if "\n" in content and len(content.split('\n')) > 1: 
                                object_elements_buffer.append(Preformatted(html.escape(display_content), style_code))
                            else:
                                object_elements_buffer.append(Paragraph(html.escape(display_content), style_body))
                            object_produced_renderable_content = True

                ts_elem = obj_elem.find("tableSection")
                if not object_produced_renderable_content and ts_elem is not None:
                    headers, rows = extract_tablesection_data(ts_elem)
                    if optimize_table_for_display(
                        table_name=field_name_display, headers=headers, rows=rows, available_width=available_width_l,
                        elements_list=object_elements_buffer,
                        default_header_style=[('BACKGROUND',(0,0),(-1,0),colors.cadetblue),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)],
                        transposed_table_new_header_row_style=[('BACKGROUND',(0,0),(-1,0),colors.cadetblue),('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke)]
                    ): object_produced_renderable_content = True

                hd_elem = obj_elem.find("hierarchyData")
                if not object_produced_renderable_content and hd_elem is not None:
                    hierarchy_tables = extract_hierarchy_data_tables(hd_elem)
                    temp_hd_sub_elements = []
                    any_hd_table_rendered = False
                    if hierarchy_tables:
                        for h_table_idx, h_table in enumerate(hierarchy_tables):
                            h_table_specific_elements = [] 
                            if optimize_table_for_display(
                                table_name=h_table['name'], headers=h_table["headers"], rows=h_table["rows"],
                                available_width=available_width_l, elements_list=h_table_specific_elements, 
                                default_header_style=[('BACKGROUND',(0,0),(-1,0),colors.lightgrey)], 
                                transposed_table_new_header_row_style=[('BACKGROUND',(0,0),(-1,0),colors.lightgrey)]
                            ):
                                any_hd_table_rendered = True
                                temp_hd_sub_elements.append(Paragraph(f"Data Table: {html.escape(h_table['name'])}", style_h4_table_title))
                                temp_hd_sub_elements.extend(h_table_specific_elements) 
                    if any_hd_table_rendered:
                        object_elements_buffer.extend(temp_hd_sub_elements) 
                        object_produced_renderable_content = True

                doc_elem = obj_elem.find("document")
                if not object_produced_renderable_content and doc_elem is not None:
                    doc_type = doc_elem.get("documentType")
                    b64_content = doc_elem.text 

                    if b64_content and b64_content.strip():
                        if doc_type == "1": 
                            extracted_word_text = extract_word_text_from_base64(b64_content.strip())
                            if extracted_word_text and extracted_word_text.strip() and \
                               not extracted_word_text.lower().startswith("skipped word document:") and \
                               not extracted_word_text.lower().startswith("error extracting text"):
                                object_elements_buffer.append(Paragraph(f"Content from Word Document ({html.escape(field_name_display)}):", style_h4_table_title))
                                display_word_text = (extracted_word_text[:3000] + '...') if len(extracted_word_text) > 3000 else extracted_word_text
                                object_elements_buffer.append(Preformatted(html.escape(display_word_text), style_code))
                                object_produced_renderable_content = True
                            elif extracted_word_text: 
                                object_elements_buffer.append(Paragraph(extracted_word_text, style_note)) 
                                object_produced_renderable_content = True

                        elif doc_type == "2": 
                            excel_sheets_data_list = extract_excel_data_from_base64(b64_content.strip())

                            temp_elements_for_this_excel_object = []
                            excel_content_was_rendered = False

                            for sheet_data_container in excel_sheets_data_list:
                                sheet_name = sheet_data_container["sheet_name"]
                                tables_in_sheet = sheet_data_container.get("tables", [])

                                if not tables_in_sheet:

                                    if sheet_name == "Error Processing Excel File" and tables_in_sheet and len(tables_in_sheet) > 0: 
                                        error_table = tables_in_sheet[0] 
                                        temp_elements_for_this_excel_object.append(Paragraph(f"Error in Excel attachment '{html.escape(field_name_display)}':", style_error))
                                        if error_table.get("rows") and error_table["rows"] and error_table["rows"][0]:

                                            error_detail = error_table["rows"][0][0] if error_table["rows"][0] else "Unknown error detail"
                                            temp_elements_for_this_excel_object.append(Paragraph(html.escape(str(error_detail)), style_body_small))
                                        excel_content_was_rendered = True 
                                    continue 

                                temp_elements_for_this_excel_object.append(Paragraph(f"Data from Sheet: '{html.escape(sheet_name)}'", style_h4_table_title)) 
                                temp_elements_for_this_excel_object.append(Spacer(1, 0.05 * inch))

                                for table_idx, current_table_obj in enumerate(tables_in_sheet):

                                    table_title = current_table_obj.get("table_title", f"Table {table_idx + 1}")
                                    table_headers = current_table_obj.get("headers", [])
                                    table_rows = current_table_obj.get("rows", [])

                                    if not table_headers and not table_rows: 
                                        continue

                                    single_table_elements_buffer = [] 

                                    if table_title and not table_title.lower().startswith("table "): 
                                        temp_elements_for_this_excel_object.append(Paragraph(f"{html.escape(table_title)}:", style_body_small)) 

                                    if optimize_table_for_display(
                                        table_name=table_title, 
                                        headers=table_headers,
                                        rows=table_rows,
                                        available_width=available_width_l,
                                        elements_list=single_table_elements_buffer, 
                                        default_header_style=[('BACKGROUND', (0, 0), (-1, 0), colors.palegreen), ('TEXTCOLOR',(0,0),(-1,0),colors.black)],
                                        transposed_table_new_header_row_style=[('BACKGROUND', (0, 0), (-1, 0), colors.palegreen), ('TEXTCOLOR',(0,0),(-1,0),colors.black)]
                                    ):
                                        temp_elements_for_this_excel_object.extend(single_table_elements_buffer)
                                        temp_elements_for_this_excel_object.append(Spacer(1, 0.1 * inch)) 
                                        excel_content_was_rendered = True

                                if tables_in_sheet and excel_content_was_rendered : 

                                    pass 

                            if excel_content_was_rendered and tables_in_sheet: 
                                temp_elements_for_this_excel_object.append(Spacer(1, 0.15 * inch)) 

                            if excel_content_was_rendered:

                                object_elements_buffer.extend(temp_elements_for_this_excel_object)
                                object_produced_renderable_content = True

                addin_elem = obj_elem.find("addin")
                if not object_produced_renderable_content and addin_elem is not None:
                    addin_data = addin_elem.get("data", "")
                    if addin_data and addin_data.strip():
                        display_addin = (addin_data[:2000] + '...') if len(addin_data) > 2000 else addin_data
                        object_elements_buffer.append(Preformatted(html.escape(display_addin), style_code)) 
                        object_produced_renderable_content = True

                if not object_produced_renderable_content:
                    temp_fallback_elements = []
                    is_fallback_content = False
                    ancillary_data_elem = obj_elem.find("ancillaryData")
                    if ancillary_data_elem is not None and ancillary_data_elem.get('extension'): 
                        temp_fallback_elements.append(Paragraph(f"Note: Ancillary data found with extension '{html.escape(ancillary_data_elem.get('extension', 'N/A'))}' for field '{field_name_display}'. Specific rendering not implemented.", style_body_small))
                        is_fallback_content = True
                    if is_fallback_content:
                        object_elements_buffer.extend(temp_fallback_elements)
                        object_produced_renderable_content = True 

                if object_produced_renderable_content:
                    if field_name_display and field_name_display != f"Unnamed Field {obj_idx + 1}" and object_elements_buffer:
                        current_section_elements_buffer.append(Paragraph(field_name_display, style_h3))
                    current_section_elements_buffer.extend(object_elements_buffer)
                    current_section_elements_buffer.append(Spacer(1, 0.05*inch)) 
                    section_has_renderable_content = True

            if section_has_renderable_content:
                final_elements_for_pdf.append(Paragraph(sec_name, style_h2)) 
                final_elements_for_pdf.extend(current_section_elements_buffer)
                if section_idx < len(all_sections) - 1: 
                    final_elements_for_pdf.append(PageBreak())
    try:
        if not final_elements_for_pdf or \
           (len(final_elements_for_pdf) == 2 and isinstance(final_elements_for_pdf[0], Paragraph) and isinstance(final_elements_for_pdf[1], Spacer) and (section_set_view is None or not all_sections)):
            final_elements_for_pdf = [Paragraph("No displayable content found in the XML after processing.", style_body)]
        elif final_elements_for_pdf and isinstance(final_elements_for_pdf[-1], PageBreak): 
            final_elements_for_pdf.pop()
        doc.build(final_elements_for_pdf,
                  onFirstPage=_add_logo_on_every_page,
                  onLaterPages=_add_logo_on_every_page)
        return pdf_buffer.getvalue()
    except Exception as e:
        print(f"PDF ERR: Final PDF building failed: {e}")
        import traceback; traceback.print_exc()
        error_buffer_final = io.BytesIO()
        try:
            error_doc_elements = [Paragraph("FATAL ERROR: Could not build PDF.", style_h1), Paragraph(str(e), style_error)]
            error_doc_final = SimpleDocTemplate(error_buffer_final, pagesize=landscape(letter))
            error_doc_final.build(error_doc_elements) 
            return error_buffer_final.getvalue()
        except Exception as final_err_build:
            print(f"ERR: Could not even build the final error PDF: {final_err_build}")
            return None

def main():
    print("--- XML to PDF Converter ---")

    try:
        SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    except NameError: 
        SCRIPT_DIR = os.getcwd() 
        print(f"WARN: __file__ not defined, using current working directory for SCRIPT_DIR: {SCRIPT_DIR}")

    SCHEMA_FOLDER = os.path.join(SCRIPT_DIR, "schemas") 

    SCHEMA_FILES_CONFIG = {
        "media_feed_schema": os.path.join(SCHEMA_FOLDER, "Media Feed Reagent Solution.xml"),
        "fed_batch_schema": os.path.join(SCHEMA_FOLDER, "FedBatch Conditions-DataCollection-CBD-UBD.xml"),
        "pa_feedback_output_schema": os.path.join(SCHEMA_FOLDER, "PA Feedback-Output.xml"),
        "pa_feedback_input_schema": os.path.join(SCHEMA_FOLDER, "PA Feedback-Input.xml"),
        "dbc_schema": os.path.join(SCHEMA_FOLDER, "DBC.xml"),
        "dbd_buffer_prep_schema": os.path.join(SCHEMA_FOLDER, "DBD Buffer Prep.xml"),
        "eluate_frac_pool_prep2_schema": os.path.join(SCHEMA_FOLDER, "Eluate Frac Pool Prep2.xml"),
        "ft_and_ht_schema": os.path.join(SCHEMA_FOLDER, "FT and HT.xml"),
        "output_schema_definition": os.path.join(SCHEMA_FOLDER, "Output.xml"),
        "tff_process_schema": os.path.join(SCHEMA_FOLDER, "TFF Process.xml"),

    }
    print(f"Expecting schema files to be in: {SCHEMA_FOLDER}")
    print(f"Predefined schemas to load: {list(SCHEMA_FILES_CONFIG.keys())}")

    set_pdf_logo_path(None)
    print("No logo will be used (default).")

    xml_file_path = ""
    while not xml_file_path:
        path = input("Enter the full path to your input XML file: ").strip()
        if os.path.isfile(path) and path.lower().endswith(".xml"):
            xml_file_path = path
        else:
            print("Invalid path or not an XML file. Please try again.")

    print(f"Reading XML file: {xml_file_path}")
    xml_content = ""
    try:
        with open(xml_file_path, 'r', encoding='utf-8') as f:
            xml_content = f.read()
        print("XML content read successfully.")
    except Exception as e:
        print(f"ERR: Could not read XML file '{xml_file_path}': {e}")
        sys.exit(1) 

    if not xml_content:
        print("ERR: XML content is empty. Cannot proceed.")
        sys.exit(1)

    print("Processing XML and generating PDF...")
    pdf_bytes = process_xml_to_pdf(xml_content, SCHEMA_FILES_CONFIG)

    if pdf_bytes:
        output_pdf_path = os.path.splitext(xml_file_path)[0] + "_output.pdf"
        try:
            with open(output_pdf_path, 'wb') as f:
                f.write(pdf_bytes)
            print(f"PDF generated successfully: {output_pdf_path}")
        except Exception as e:
            print(f"ERR: Could not save PDF to '{output_pdf_path}': {e}")
    else:
        print("ERR: PDF generation failed. No PDF bytes returned.")

if __name__ == "__main__":
    main()